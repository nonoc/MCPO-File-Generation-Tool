"""
Excel (XLSX) exporter module using openpyxl.
"""

import os
import logging
from typing import Any, List, Optional, Dict

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Alignment, PatternFill

# Import from new modules (try relative first)
try:
    from ..shared.utils import _generate_unique_folder, _public_url
    from ..shared.constants import XLSX_TEMPLATE, XLSX_TEMPLATE_PATH
    from ..processing.text_utils import _normalize_table_rows
    from ..processing.block_utils import _convert_markdown_to_structured
except ImportError:
    from shared.utils import _generate_unique_folder, _public_url
    from shared.constants import XLSX_TEMPLATE, XLSX_TEMPLATE_PATH
    from processing.text_utils import _normalize_table_rows
    from processing.block_utils import _convert_markdown_to_structured

log = logging.getLogger(__name__)


def _find_table_in_blocks(blocks: Any) -> Optional[List[List[str]]]:
    """
    Recursively search for table data in nested block structures.
    
    Searches top-level blocks for table type, then recursively searches
    children arrays and items arrays.
    
    Args:
        blocks: Content blocks (dict, list, or other)
        
    Returns:
        Normalized table data if found, None otherwise
    """
    if not blocks:
        return None
    
    # If it's a single block dict, check if it's a table
    if isinstance(blocks, dict):
        block_type = (blocks.get("type") or "").strip().lower()
        if block_type == "table":
            data = blocks.get("data") or blocks.get("rows") or blocks.get("cells") or []
            return _normalize_table_rows(data)
        
        # Recursively search children arrays
        for key in ("children", "content", "blocks"):
            child_value = blocks.get(key)
            if child_value:
                result = _find_table_in_blocks(child_value)
                if result:
                    return result
        
        # Recursively search items arrays
        items = blocks.get("items")
        if items:
            result = _find_table_in_blocks(items)
            if result:
                return result
        
        return None
    
    # If it's a list of blocks, search each one
    if isinstance(blocks, list):
        for block in blocks:
            result = _find_table_in_blocks(block)
            if result:
                return result
        return None
    
    return None


def _render_structured_content_to_excel(ws, data: List[Dict], start_row: int = 1, start_col: int = 1) -> int:
    """
    Render structured blocks to Excel worksheet.
    
    Renders blocks in order:
    - Title: first row, bold
    - Paragraphs: each to new row
    - List items: each to new row
    - Table: written as grid starting after other content
    
    Args:
        ws: OpenPyXL worksheet
        data: List of structured block dictionaries
        start_row: Starting row (1-indexed)
        start_col: Starting column (1-indexed)
        
    Returns:
        Next available row after rendering
    """
    current_row = start_row
    
    for block in data:
        block_type = (block.get("type") or "").strip().lower()
        
        if block_type == "title":
            # Write title to first row, bold
            title_text = block.get("text") or block.get("title") or ""
            cell = ws.cell(row=current_row, column=start_col)
            cell.value = title_text
            cell.font = Font(bold=True)
            current_row += 1
            
        elif block_type == "paragraph":
            # Write paragraph to new row
            text = block.get("text") or ""
            cell = ws.cell(row=current_row, column=start_col)
            cell.value = text
            current_row += 1
            
        elif block_type in ("bullet", "list"):
            # Write list items
            items = block.get("items") or []
            if items:
                # Handle items list format: {"type": "list", "items": [...]}
                for item in items:
                    text = item.get("text") if isinstance(item, dict) else item
                    cell = ws.cell(row=current_row, column=start_col)
                    cell.value = text
                    current_row += 1
            elif block.get("text"):
                # Handle direct text format: {"type": "bullet", "text": "item"}
                text = block.get("text")
                cell = ws.cell(row=current_row, column=start_col)
                cell.value = text
                current_row += 1
                
        elif block_type == "table":
            # Write table as grid
            table_data = _normalize_table_rows(
                block.get("data") or block.get("rows") or block.get("cells") or []
            )
            for row_idx, row in enumerate(table_data):
                for col_idx, cell_value in enumerate(row):
                    ws.cell(
                        row=current_row + row_idx,
                        column=start_col + col_idx
                    ).value = cell_value
            current_row += len(table_data)
    
    return current_row


def _preprocess_excel_content(content: Any) -> List[List[str]]:
    """
    Preprocess content to normalize to 2D list format for Excel.
    
    Handles:
    - String content (markdown or plain text)
    - Dict content (structured blocks)
    - List content (mixed blocks)
    - Already normalized 2D list (returned as-is)
    
    Args:
        content: Content in various formats
        
    Returns:
        Normalized 2D list of cell values
    """
    if content is None:
        return []
    
    # If already a 2D list, return as-is (with normalized rows)
    if isinstance(content, list) and content:
        # Check if it's a 2D list (list of lists)
        if isinstance(content[0], list):
            # Return 2D list as-is (already in Excel format)
            return content
        # It's a list of structured blocks - use recursive search to find tables
        table_data = _find_table_in_blocks(content)
        if table_data:
            return table_data
        # No table blocks found, try to normalize as rows
        return _normalize_table_rows(content)
    
    if isinstance(content, str):
        # Parse markdown to structured blocks
        structured = _convert_markdown_to_structured(content)
        if not structured:
            return []
        # Extract tables from structured content
        tables = []
        for block in structured:
            if block.get("type") == "table" and block.get("data"):
                tables.append(_normalize_table_rows(block.get("data")))
        if tables:
            # Return first table found, or flatten content if no tables
            return tables[0] if tables else []
        # No tables found, return content as single cell
        return [[content]]
    
    if isinstance(content, dict):
        # Single dict - check if it's a table
        content_type = (content.get("type") or "").strip().lower()
        if content_type == "table":
            data = content.get("data") or content.get("rows") or content.get("cells") or []
            return _normalize_table_rows(data)
        # Otherwise, extract text content
        text = content.get("text") or content.get("title") or content.get("content") or ""
        if text:
            return [[str(text)]]
        return []
    
    return []


def _create_excel(data: Any, filename: str, folder_path: Optional[str] = None, title: Optional[str] = None) -> dict:
    """
    Create an Excel file from data.
    
    Args:
        data: Content in various formats (string, dict, list, or 2D list)
        filename: Output filename
        folder_path: Optional folder path for output
        title: Optional sheet title
        
    Returns:
        Dict with url and path to generated file
    """
    log.debug("Creating Excel file with optional template")
    
    if folder_path is None:
        folder_path = _generate_unique_folder()
    
    if filename:
        filepath = os.path.join(folder_path, filename)
        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        fname = filename
    else:
        from ..shared.utils import _generate_filename
        filepath, fname = _generate_filename(folder_path, "xlsx")

    # Detect structured content (list of dicts with 'type' field) BEFORE preprocessing
    # This allows us to render full structured content instead of just extracting tables
    is_structured = (
        isinstance(data, list)
        and data
        and isinstance(data[0], dict)
        and "type" in data[0]
    )
    
    # Detect if this is raw 2D matrix data (clean spreadsheet mode)
    # A 2D matrix has all rows as lists of primitives (strings, numbers, booleans, None)
    is_raw_matrix = (
        isinstance(data, list)
        and data
        and isinstance(data[0], list)
        and all(
            isinstance(cell, (str, int, float, bool, type(None)))
            for row in data[:5]  # Sample check first 5 rows
            for cell in row[:10]  # Sample check first 10 cells per row
        )
    )
    
    # For raw 2D matrices: create clean workbook
    # For structured content: create clean workbook (no template needed for spreadsheet mode)
    if is_raw_matrix or is_structured:
        log.debug("Raw 2D matrix or structured content detected, creating clean workbook")
        wb = Workbook()
    elif XLSX_TEMPLATE and XLSX_TEMPLATE_PATH:
        try:
            log.debug("Loading XLSX template...")
            wb = load_workbook(XLSX_TEMPLATE_PATH)
            log.debug(f"Template loaded with {len(wb.sheetnames)} sheet(s)")
        except Exception as e:
            log.warning(f"Failed to load XLSX template: {e}")
            wb = Workbook()
    else:
        log.debug("No XLSX template available, creating new workbook")
        wb = Workbook()

    ws = wb.active

    # Set sheet title if provided (applies to all modes)
    if title:
        ws.title = "".join(c for c in title if c.isalnum() or c in (' ', '-', '_')).strip()[:31]

    # Handle structured content differently from raw matrices
    if is_structured:
        # Render structured blocks directly to Excel
        log.debug("Rendering structured content to Excel")
        _render_structured_content_to_excel(ws, data)
    elif is_raw_matrix:
        # Raw 2D matrix: use simplified write logic
        log.debug("Writing raw 2D matrix to Excel")
        start_row, start_col = 1, 1
        
        # Handle empty data
        if not data:
            wb.save(filepath)
            return {"url": _public_url(folder_path, fname), "path": filepath}
        
        # Write raw matrix at A1
        max_cols = max(len(row) for row in data) if data else 0
        for r, row in enumerate(data):
            for c, cell_value in enumerate(row):
                cell = ws.cell(row=start_row + r, column=start_col + c)
                cell.value = cell_value
                if r == 0 and cell_value:  # Header row
                    cell.font = Font(bold=True)
        
        # Auto-adjust column widths
        for c in range(max_cols):
            max_len = 0
            for r in range(len(data)):
                if c < len(data[r]):
                    max_len = max(max_len, len(str(data[r][c])))
            ws.column_dimensions[get_column_letter(start_col + c)].width = min(max_len + 2, 150)
    else:
        # Template mode: preprocess content and use template logic
        log.debug("Using template mode for content")
        data = _preprocess_excel_content(data)
        
        # Try to find and replace title cell
        if title:
            title_cell_found = False
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and "title" in cell.value.lower():
                        cell.value = title
                        cell.font = Font(bold=True)
                        log.debug(f"Title '{title}' replaced in cell {get_column_letter(cell.column)}{cell.row} containing 'title'")
                        title_cell_found = True
                        break
                if title_cell_found:
                    break
        
        # Detect auto_filter position for data placement
        start_row, start_col = 1, 1
        if ws.auto_filter and ws.auto_filter.ref:
            try:
                from openpyxl.utils import range_boundaries
                start_col, start_row, _, _ = range_boundaries(ws.auto_filter.ref)
            except Exception:
                pass

        # Handle empty data
        if not data:
            wb.save(filepath)
            return {"success": True, "filepath": filepath, "filename": filename}

        # Get template border style
        template_border = ws.cell(start_row, start_col).border
        has_borders = template_border and any([
            template_border.top.style, 
            template_border.bottom.style, 
            template_border.left.style, 
            template_border.right.style
        ])
        
        # Write data
        max_cols = max(len(row) for row in data) if data else 0
        for r in range(max(len(data) + 10, 50)):
            for c in range(max(max_cols + 5, 20)):
                cell = ws.cell(row=start_row + r, column=start_col + c)
                
                if r < len(data) and c < max_cols:
                    if c < len(data[r]):
                        cell.value = data[r][c]
                    else:
                        cell.value = None
                    if r == 0 and cell.value:  # Header row
                        cell.font = Font(bold=True)
                    if has_borders:  # Apply border if template has it
                        cell.border = Border(
                            top=template_border.top, 
                            bottom=template_border.bottom,
                            left=template_border.left, 
                            right=template_border.right
                        )
                else:
                    cell.value = None
                    if cell.has_style:
                        cell.font = Font()
                        cell.fill = PatternFill()
                        cell.border = Border()
                        cell.alignment = Alignment()

        # Update auto-filter range if present (only if there are columns)
        if ws.auto_filter and max_cols > 0:
            ws.auto_filter.ref = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(start_col + max_cols - 1)}{start_row + len(data) - 1}"
        
        # Auto-adjust column widths
        for c in range(max_cols):
            max_len = 0
            for r in range(len(data)):
                if c < len(data[r]):
                    max_len = max(max_len, len(str(data[r][c])))
            ws.column_dimensions[get_column_letter(start_col + c)].width = min(max_len + 2, 150)

    wb.save(filepath)
    return {"url": _public_url(folder_path, fname), "path": filepath}


def add_auto_sized_review_comment(cell, text: str, author: str = "AI Reviewer") -> None:
    """
    Add a comment to a cell with auto-sizing.
    
    Args:
        cell: OpenPyXL cell object
        text: Comment text
        author: Comment author
    """
    if not text:
        return

    import math
    avg_char_width = 7
    px_per_line = 15
    base_width = 200
    max_width = 500
    min_height = 40

    width = min(max_width, base_width + len(text) * 2)
    chars_per_line = max(1, width // avg_char_width)
    lines = 0
    for paragraph in text.split('\n'):
        lines += math.ceil(len(paragraph) / chars_per_line)
    height = max(min_height, lines * px_per_line)

    comment = cell.comment if cell.comment else cell._comment
    if comment is None:
        from openpyxl.comments import Comment
        comment = Comment(text, author)
        comment.width = width
        comment.height = height
        cell.comment = comment
    else:
        comment.text = text
        comment.author = author
        comment.width = width
        comment.height = height
